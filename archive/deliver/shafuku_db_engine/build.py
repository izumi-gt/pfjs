# -*- coding: utf-8 -*-
"""
DB構築（Excel出力）。新規作成 or 既存DBへの追記に対応。
出力後に外部キー違反・fact主キー重複・local親参照を自己検査する。
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

from . import schema
from .masters import (CF_MASTER, PL_MASTER, BS_MASTER,
                      ACT_FULL_CF, ACT_FULL_PL, SEC_NAME_BS)

FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=9)
LABEL = Font(name=FONT, size=9)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
RIGHT = Alignment(horizontal="right")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
NUMFMT = '#,##0;(#,##0);"-"'


def global_account_rows():
    rows = []
    for c, act, io, nm, tot in CF_MASTER:
        rows.append([c, "資金収支計算書", ACT_FULL_CF.get(act, act), io, nm, tot])
    for c, act, io, nm, tot in PL_MASTER:
        rows.append([c, "事業活動計算書", ACT_FULL_PL.get(act, act), io, nm, tot])
    for m in BS_MASTER:
        c, sec, grp, nm, tot = m[0], m[1], m[2], m[3], m[4]
        rows.append([c, "貸借対照表", SEC_NAME_BS.get(sec, sec), grp, nm, tot])
    return rows


def _sheet(wb, name, headers, rows, num_cols=()):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CTR; c.border = BORDER
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v); c.font = LABEL; c.border = BORDER
            if j in num_cols:
                c.number_format = NUMFMT; c.alignment = RIGHT
    ws.freeze_panes = "A2"
    for j, h in enumerate(headers, 1):
        vals = [len(str(h))] + [len(str(r[j-1])) for r in rows if j-1 < len(r) and r[j-1] is not None]
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(42, max(vals or [10]) + 2))
    return ws


def build(ingested_list, out_path, append_to=None):
    """
    ingested_list: list[Ingested]（複数法人可）
    append_to: 既存DBパス。指定時はその fact/dim_corp/dim_segment/dim_account_local に追記。
    """
    # collect all rows
    fact = []
    corp_rows = []
    seg_rows = []
    local_rows = []
    concept_rows = []
    for ing in ingested_list:
        fact.extend(ing.fact)
        corp_rows.append(ing.corp_row)
        seg_rows.extend(ing.segments)
        local_rows.extend(ing.locals)
        concept_rows.extend(getattr(ing, "concepts", []))

    if append_to:
        prev = _read_existing(append_to)
        fact = prev["fact"] + fact
        corp_rows = prev["corp"] + corp_rows
        seg_rows = prev["seg"] + seg_rows
        local_rows = prev["local"] + local_rows

    # ---- post-checks ----
    problems = post_checks(fact, seg_rows, local_rows)

    # ---- write ----
    wb = Workbook(); wb.remove(wb.active)
    _readme(wb, corp_rows)
    _sheet(wb, "fact_financial", schema.TABLES["fact_financial"], fact, num_cols=(7,))
    _sheet(wb, "dim_corp", schema.TABLES["dim_corp"], corp_rows)
    _sheet(wb, "dim_form", schema.TABLES["dim_form"], [list(f) for f in schema.FORMS])
    _sheet(wb, "dim_account_global", schema.TABLES["dim_account_global"], global_account_rows())
    _sheet(wb, "dim_account_local", schema.TABLES["dim_account_local"], local_rows)
    _sheet(wb, "dim_account_concept", schema.TABLES["dim_account_concept"], _dedup(concept_rows))
    _sheet(wb, "dim_segment", schema.TABLES["dim_segment"], seg_rows)
    order = ["README", "fact_financial", "dim_corp", "dim_form",
             "dim_account_global", "dim_account_local", "dim_account_concept", "dim_segment"]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.save(out_path)
    return problems, len(fact)


def post_checks(fact, seg_rows, local_rows):
    e = []
    gcodes = {r[0] for r in global_account_rows()}
    lcodes = {r[0] for r in local_rows}
    scodes = {r[0] for r in seg_rows}
    fcodes = {f[0] for f in schema.FORMS}
    allc = gcodes | lcodes
    for f in fact:
        if f[3] not in allc: e.append(f"FK科目: {f[3]} 不明")
        if f[4] not in scodes: e.append(f"FK事業区分: {f[4]} 不明")
        if f[2] not in fcodes: e.append(f"FK様式: {f[2]} 不明")
    for r in local_rows:
        # parent が空文字 or 拠点擬似親(@LOC:...) は「最上位＝親なし」の正当な表現。
        # マスタ未収載の事業別大区分を拠点別 local として採番した場合に生じる。
        parent = r[2]
        if parent in ("", None) or str(parent).startswith("@LOC:"):
            continue
        if parent not in allc: e.append(f"FKlocal親: {parent} 不明")
    pk = [tuple(f[i] for i in range(6)) for f in fact]  # 法人,年度,様式,科目,区分,metric
    dups = [k for k, c in Counter(pk).items() if c > 1]
    for k in dups[:20]:
        e.append(f"fact主キー重複: {k}")
    # dedupe within report (unique)
    seen = set(); uniq = []
    for x in e:
        if x not in seen: seen.add(x); uniq.append(x)
    return uniq


def _read_existing(path):
    wb = load_workbook(path, read_only=True)
    def rows(name):
        ws = wb[name]; rs = list(ws.iter_rows(values_only=True))
        return [list(r) for r in rs[1:] if any(x is not None for x in r)]
    out = {"fact": rows("fact_financial"), "corp": rows("dim_corp"),
           "seg": rows("dim_segment"), "local": rows("dim_account_local")}
    try:
        out["concept"] = rows("dim_account_concept")
    except KeyError:
        out["concept"] = []
    return out


def _readme(wb, corp_rows):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    lines = [
        "決算情報データベース（社会福祉法人 決算情報 統合DB）",
        "",
        "■ 設計思想（SQL移行前提・1シート=1テーブル）",
        "・fact_financial = 縦持ち事実表。金額のみ保持し属性は dim に正規化。",
        "・科目は2層: global(全国共通の幹) と local(法人別の小区分=枝、親コード付き)。",
        "・法人横断比較は global科目コードで突合。明細は dim_account_local.親科目コード で幹に紐付く。",
        "・local科目コードは (拠点, 親, 名称) で一意化（同名小区分が別の親に並存しても衝突しない）。",
        "",
        "■ テーブル",
        "・fact_financial / dim_corp / dim_form / dim_account_global / dim_account_local / dim_segment",
        "",
        "■ 検算（取り込み前に validate.validate で全件自己検証。不一致が1件でもあれば中止）",
        "・CF縦計・収支差額式 / PL活動区分連鎖 / BSバランス / 2階層 小区分=親 /",
        "  区分和=合計 / 合計-内部消去=法人合計 / 様式間突合 / 計算書またぎ(BS↔PL)。",
        "・出力後に FK違反・fact主キー重複・local親参照を post_checks で再検査。",
        "",
        "■ 収録法人",
    ]
    for r in corp_rows:
        lines.append(f"・{r[1]}（{r[0]}）{r[3]} / 拠点{r[4]}")
    for i, t in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=t).font = Font(name=FONT, bold=(i == 1 or t.startswith("■")), size=(12 if i == 1 else 9))
    ws.column_dimensions["A"].width = 110


def _dedup(rows):
    seen=set(); out=[]
    for r in rows:
        k=tuple(r)
        if k not in seen:
            seen.add(k); out.append(r)
    return out
